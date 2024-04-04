import openpyxl as oxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
import numpy as np
from datetime import datetime

def GetDataByPerson(totalData, person):
    data = []
    for i in range(len(totalData)):
        if totalData[i][1] == person:
            data.append(totalData[i])
    return data

def FormatDate(data, date_colNum = 0):
    for i in range(len(data)):
        data[i][date_colNum] = data[i][date_colNum].strftime("%Y-%m-%d")
        data[i][date_colNum] = data[i][date_colNum][:4] + "년 " + data[i][date_colNum][5:7] + "월 " + data[i][date_colNum][-1] + "주차"
    return data

def GetDatesCount(data, date_colNum = 0):
    dates = []
    date_counts = []
    for i in range(len(data)):
        if data[i][date_colNum] not in dates:
            dates.append(data[i][date_colNum])
            date_counts.append(0)
        else:
            date_counts[dates.index(data[i][date_colNum])] += 1
    return dates, date_counts

def MergeWriteCells(sheet, start_row, start_col, end_row, end_col, value):
    sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    sheet.cell(row=start_row, column=start_col, value=value)
    sheet.cell(row=start_row, column=start_col).alignment = Alignment(horizontal="center", vertical="center")    # align the text to the center

def WriteByCategory(sheet, data, categories, time_colNum):    
    for i in range(len(data)):
        for j in range(len(categories)):
            if data[i][category_colNum] == categories[j]:
                sheet.cell(row=2+date_counts[dates.index(data[i][0])], column=2*j+1, value=data[i][-1])
                sheet.cell(row=2+date_counts[dates.index(data[i][0])], column=2*j+1).alignment = Alignment(horizontal="center", vertical="center")

def OutlineBox(sheet, start_row, start_col, end_row, end_col, style='thin'):
    for i in range(start_row, end_row+1):
        for j in range(start_col, end_col+1):
            cell = sheet.cell(row=i, column=j)
            if i == start_row:
                if j == start_col:
                    cell.border = oxl.styles.Border(top=oxl.styles.Side(style), left=oxl.styles.Side(style))
                elif j == end_col:
                    cell.border = oxl.styles.Border(top=oxl.styles.Side(style), right=oxl.styles.Side(style))
                else:
                    cell.border = oxl.styles.Border(top=oxl.styles.Side(style))
            elif i == end_row:
                if j == start_col:
                    cell.border = oxl.styles.Border(bottom=oxl.styles.Side(style), left=oxl.styles.Side(style))
                elif j == end_col:
                    cell.border = oxl.styles.Border(bottom=oxl.styles.Side(style), right=oxl.styles.Side(style))
                else:
                    cell.border = oxl.styles.Border(bottom=oxl.styles.Side(style))
            else:
                if j == start_col:
                    cell.border = oxl.styles.Border(left=oxl.styles.Side(style))
                elif j == end_col:
                    cell.border = oxl.styles.Border(right=oxl.styles.Side(style))


person = "김다은"
name_colNum = 1
content_colNum = 3
tag_colNum = 5
category_colNum = 4
categoires = ["프로젝트", "공부", "취준", "기타"]
time_colNum = len(categoires) * 2 + 1
total_sheetName = "전체 한 일"
weeklyCheck_sheetName = "주간 점검"
colorPalette = [
  "FCEBB6",
  "FFD8D8",
  "FFC3A0",
  "FFDAC1",
  "FFB7B2",
  "FFD8CC",
  "FFB6C1",
  "FFC0CB",
  "F8BBD0",
  "E1BEE7",
  "D1C4E9",
  "C5CAE9",
  "BBDEFB",
  "B3E5FC",
  "B2EBF2",
  "B2DFDB",
  "C8E6C9",
  "DCEDC8",
  "F0F4C3",
  "FFF9C4",
  "FFECB3",
  "FFE0B2",
  "FFCCBC",
  "FFAB91"
]
dateColorPalette = [
    ["C7D8ED", "F4B3C2", "F2E0B7", "A3D6C1", "E6CF8B"], # January
    ["E7CCE2", "FFD1DC", "F1B9CE", "A9D0F5", "F5E1A8"], # February
    ["C2E0C4", "F9D6E1", "BBD7F1", "FFE8B4", "D8CCEB"], # March
    ["F8E5E5", "AED6F1", "FFF4CF", "C7E2C9", "FAD4A0"], # April
    ["FFDDC1", "C2D6E2", "F3D9E2", "D1E6C9", "F5E1A8"], # May
    ["FFDFBA", "A5D8E2", "F4CFD4", "D1E6C9", "FFD7C1"], # June
    ["FFC2C2", "FFDFBA", "B5D8E2", "FFE8B4", "C7E2C9"], # July
    ["FFD1DC", "BBD7F1", "F3D9E2", "C7E2C9", "FFD7C1"], # August
    ["F3D9E2", "AED6F1", "FFF4CF", "C2E0C4", "FAD4A0"], # September
    ["F5E1A8", "FFC2C2", "D1E6C9", "F4CFD4", "FFDFBA"], # October
    ["C7D8ED", "F4B3C2", "F2E0B7", "A3D6C1", "E6CF8B"], # November
    ["F4B3C2", "C7D8ED", "F2E0B7", "A3D6C1", "E6CF8B"]  # December
]


# create workbook and load the workbook
wb = oxl.Workbook()
lwb = oxl.load_workbook("Data.xlsx")
sheet = lwb[total_sheetName]

# get the data from the sheet A to G which has B column 
total_data = []
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=7):
    total_data.append([cell.value for cell in row])

data = GetDataByPerson(total_data, person)
data = FormatDate(data)

# create new excel file named "TimeLine.xlsx"
lwb = oxl.Workbook()
sheet = lwb.create_sheet("timeline_" + person)

# write first row
sheet.cell(row=1, column=time_colNum, value="날짜")
for i in range(len(categoires)):
    sheet.merge_cells(start_row=1, start_column=2*i+1, end_row=1, end_column=2*i+2)
    sheet.cell(row=1, column=2*i+1, value=categoires[i])
    sheet.cell(row=1, column=2*i+1).alignment = Alignment(horizontal="center", vertical="center")    # align the text to the center

dates, date_counts = GetDatesCount(data) # get dates from the data and count the number of the same date

# sort data by date, category, and tag
data = sorted(data, key=lambda x: (x[0], x[category_colNum], x[tag_colNum]))
# delete evey 1, 2 elements in the data
data = np.delete(data, [1, 2], axis=1)
# swap 1 <-> 3 elements in the data
data[:, [1, 3]] = data[:, [3, 1]]
# swap 1 <-> 2 elements in the data
data[:, [1, 2]] = data[:, [2, 1]]

# collect data by date and category
data_by_date_category = []
for i in range(len(dates)):
    data_by_date_category.append([])
    for j in range(len(categoires)):
        data_by_date_category[i].append([])
        for k in range(len(data)):
            if data[k][0] == dates[i] and data[k][1] == categoires[j]:
                data_by_date_category[i][j].append(data[k][2:])

# sort data by tag in the data_by_date_category. tag index is 0
for i in range(len(data_by_date_category)):
    for j in range(len(data_by_date_category[i])):
        data_by_date_category[i][j] = sorted(data_by_date_category[i][j], key=lambda x: x[0])

# make assgin color to each tag
tag_color = {}
for i in range(len(data)):
    tag = data[i][2]
    if tag not in tag_color.keys():
        tag_color[tag] = colorPalette[len(tag_color) % len(colorPalette)]

# wirte tag and content divided by category + merge tags and stylize
startRow = 2
for i in range(len(dates)):
    max_length = 0
    for j in range(len(categoires)):
        #get max length of the data
        for k in range(len(data_by_date_category[i][j])):
            max_length = max(max_length, len(data_by_date_category[i][j]))
            
            # write tag and stylize
            sheet.cell(row=startRow+k, column=2*j+1, value=data_by_date_category[i][j][k][0])
            sheet.cell(row=startRow+k, column=2*j+1).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=startRow+k, column=2*j+1).fill = oxl.styles.PatternFill(start_color=tag_color[data_by_date_category[i][j][k][0]], end_color=tag_color[data_by_date_category[i][j][k][0]], fill_type="solid") #color the cell            

            # write content and stylize
            sheet.cell(row=startRow+k, column=2*j+2, value=data_by_date_category[i][j][k][1])
            sheet.cell(row=startRow+k, column=2*j+2).alignment = Alignment(horizontal="center", vertical="center")
            
    MergeWriteCells(sheet, startRow, time_colNum, startRow+max_length-1, time_colNum, dates[i])
    # color the date cell
    month = int(dates[i][6:8]) -1
    week = int(dates[i][-3]) -1
    sheet.cell(row=startRow, column=time_colNum).fill = oxl.styles.PatternFill(start_color=dateColorPalette[month][week], end_color=dateColorPalette[month][week], fill_type="solid")
    # outline the rows up and down from startRow to startRow+max_length-1
    for k in range(1, 12):
        sheet.cell(row=startRow, column=k).border = oxl.styles.Border(top=oxl.styles.Side(style='thin'))
        sheet.cell(row=startRow+max_length-1, column=k).border = oxl.styles.Border(bottom=oxl.styles.Side(style='thin'))
        if week == 0: # make top medium border
            sheet.cell(row=startRow, column=k).border = oxl.styles.Border(top=oxl.styles.Side(style='medium'))

    startRow += max_length

# merge tags in same date and stylize


# Merge date cells and stylize


# merge tags in same date
dateStartRow = 2
for i in range(len(dates)):
    max_length = 0
    for j in range(len(categoires)):
        max_length = max(max_length, len(data_by_date_category[i][j]))
        group = []
        groupLength = []
        startRow = dateStartRow
        for k in range(len(data_by_date_category[i][j])):
            if data_by_date_category[i][j][k][0] not in group:
                group.append(data_by_date_category[i][j][k][0])
                groupLength.append(1)
            else:
                groupLength[group.index(data_by_date_category[i][j][k][0])] += 1
        for key in range(len(group)):
            sheet.merge_cells(start_row=startRow, start_column=2*j+1, end_row=startRow+groupLength[key]-1, end_column=2*j+1)
            # outline the rows up and down from startRow to startRow+groupLength[key]-1
            for k in range(startRow, startRow+groupLength[key]):
                sheet.cell(row=k, column=2*j+1).border = oxl.styles.Border(left=oxl.styles.Side(style='thin'))
                sheet.cell(row=k, column=2*j+2).border = oxl.styles.Border(right=oxl.styles.Side(style='thin'))
                if (k == startRow):
                    sheet.cell(row=k, column=2*j+1).border = oxl.styles.Border(top=oxl.styles.Side(style='thin'), left=oxl.styles.Side(style='thin'))
                    sheet.cell(row=k, column=2*j+2).border = oxl.styles.Border(top=oxl.styles.Side(style='thin'), right=oxl.styles.Side(style='thin'))
                if (k == startRow+groupLength[key]-1):
                    sheet.cell(row=k, column=2*j+1).border = oxl.styles.Border(bottom=oxl.styles.Side(style='thin'), left=oxl.styles.Side(style='thin'))
                    sheet.cell(row=k, column=2*j+2).border = oxl.styles.Border(bottom=oxl.styles.Side(style='thin'), right=oxl.styles.Side(style='thin'))
            if (startRow == startRow+groupLength[key]-1):
                sheet.cell(row=k, column=2*j+1).border = oxl.styles.Border(top=oxl.styles.Side(style='thin'), bottom=oxl.styles.Side(style='thin'), left=oxl.styles.Side(style='thin'))
                sheet.cell(row=k, column=2*j+2).border = oxl.styles.Border(top=oxl.styles.Side(style='thin'), bottom=oxl.styles.Side(style='thin'), right=oxl.styles.Side(style='thin'))
            startRow += groupLength[key]
    dateStartRow += max_length


lwb.save("TimeLine.xlsx")
