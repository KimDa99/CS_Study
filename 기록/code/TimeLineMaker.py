import openpyxl as oxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import DataBar, FormatObject, Rule, DataBarRule
import numpy as np
from datetime import datetime


import TimeBlock

person = "김다은"
dateRange = ["2024년 01월 1주차", "2024년 03월 1주차"]

doIncludeMonthlyReview = True
doIncludeWeeklyReview = True
doFilterByDate = True

total_sheetName = "전체 한 일"
weeklyCheck_sheetName = "주간 점검"
monthlyCheck_sheetName = "월말 점검"
categories_sheetName = "categories"


ColorPalette = [
    "FCEBB6","FFD8D8","FFC3A0","FFDAC1","FFB7B2","FFD8CC",
    "FFB6C1","FFC0CB","F8BBD0","E1BEE7","D1C4E9","C5CAE9",
    "BBDEFB","B3E5FC","B2EBF2","B2DFDB","C8E6C9","DCEDC8",
    "F0F4C3","FFF9C4","FFECB3","FFE0B2","FFCCBC","FFAB91"]
DateColorPalette = [
    ["C7D8ED", "F4B3C2", "F2E0B7", "A3D6C1", "E6CF8B"], # January
    ["E7CCE2", "FFD1DC", "F1B9CE", "A9D0F5", "F5E1A8"], # February
    ["C2E0C4", "F9D6E1", "BBD7F1", "FFE8B4", "D8CCEB"], # March
    ["F8E5E5", "AED6F1", "FFF4CF", "C7E2C9", "FAD4A0"], # April
    ["FFDDC1", "C2D6E2", "F3D9E2", "D1E6C9", "F5E1A8"], # May
    ["FFDFBA", "A5D8E2", "F4CFD4", "D1E6C9", "FFD7C1"], # June
    ["FFC2C2", "FFDFBA", "B5D8E2", "FFE8B4", "C7E2C9"], # July
    ["FFD1DC", "BBD7F1", "F3D9E2", "C7E2C9", "FFD7C1"], # August
    ["F3D9E2", "AED6F1", "FFF4CF", "C2E0C4", "FAD4A0"], # September
    ["F5E1A8", "FFC2C2", "D1E6C9", "F4CFD4", "FFDFBA"], # October
    ["C7D8ED", "F4B3C2", "F2E0B7", "A3D6C1", "E6CF8B"], # November
    ["F4B3C2", "C7D8ED", "F2E0B7", "A3D6C1", "E6CF8B"]  # December
]

def MakeTimeLine(person = "김다은", dateRange = ["2024년 01월 1주차", "2024년 03월 1주차"], doIncludeMonthlyReview = True, doIncludeWeeklyReview = True, doFilterByDate = True, total_sheetName = "전체 한 일", weeklyCheck_sheetName = "주간 점검", monthlyCheck_sheetName = "월말 점검", categories_sheetName = "categories", ColorPalette = ColorPalette, DateColorPalette = DateColorPalette):
    timeLine_colNum = 0
    maxCol = 0
    tagColorPalette = {}

    #create workbook and load the workbook
    wb = oxl.Workbook()
    lwb = oxl.load_workbook("Data.xlsx")

    # get the data
    sheet = lwb[categories_sheetName]
    div = GetCategories(sheet, name="분류")
    categoires = GetCategories(sheet)
    sheet = lwb[total_sheetName]
    data = GetProcessedData(sheet, categoires, person, dateRange = dateRange)

    # get the weekly review
    sheet = lwb[categories_sheetName]
    weeklyDiv = GetCategories(sheet, name="주간점검 분류")
    weeklyReview = []
    weeklyInfoName = []
    if doIncludeWeeklyReview:
        sheet = lwb[weeklyCheck_sheetName]
        weeklyReview = GetReview(sheet, person, removeCols=[1, -1], monthOnly=False)
        # get first row of the sheet to get the review category
        weeklyInfoName = weeklyReview[0]

    # get the monthly review
    sheet = lwb[categories_sheetName]
    monthlyInfoName = GetCategories(sheet, name="월점검")
    monthlyDiv = GetCategories(sheet, name="월점검 분류")
    monthlyReview = []
    monthlyCategory = []
    monthluReviewCategoryCol = ['C', 'D', 'E', 'F']
    if doIncludeMonthlyReview:
        sheet = lwb[monthlyCheck_sheetName]
        monthlyReview = GetReview(sheet, person, removeCols=[1, -1], monthOnly=True)
        # get first row of the sheet to get the review category
        firstRow = sheet[1]
        for i in range(len(monthluReviewCategoryCol)):
            # turn in do col char to int
            colIndex = ord(monthluReviewCategoryCol[i]) - ord('A')
            monthlyCategory.append(firstRow[colIndex].value)


    ### Write the data to the excel file ###
    TimeBlocks = TimeBlock.TimeBlockList(data, reviews= weeklyReview)
    maxCol = GetMaxCol(TimeBlocks, categoires, weeklyReview, monthlyReview)
    timeLine_colNum = len(categoires) * 2 + 1


    #create excel file
    lwb = oxl.Workbook()
    sheet = lwb.create_sheet(title="TimeLine_" + person)

    # write the date line
    startRow = 2
    startCol = 1
    conditionalFormatCell =[]
    for t in range(len(TimeBlocks.TimeBlocks)):
        block = TimeBlocks.TimeBlocks[t]
    # region write date
        MergeWriteCells(sheet, startRow, timeLine_colNum, startRow + block.GetLength() -1 , timeLine_colNum, block.date)
        sheet.cell(row=startRow, column=timeLine_colNum).fill = PatternFill(start_color=GetDateColor(block.date), end_color=GetDateColor(block.date), fill_type="solid")
    # endregion

    # region write the data
        # write the data
        for cat in block.data:
            tagRow = startRow
            for tag in block.data[cat]:
                # write the tag
                catCol = categoires.index(cat) * 2 + 1
                MergeWriteCells(sheet, tagRow, catCol, tagRow + len(block.data[cat][tag]) - 1, catCol, tag)
                sheet.cell(row=tagRow, column=catCol).fill = PatternFill(start_color=GetTagColor(tag), end_color=GetTagColor(tag), fill_type="solid")

                # write the content
                cntRow = tagRow
                for content in block.data[cat][tag]:
                    sheet.cell(row=cntRow, column=catCol + 1, value=content)
                    cntRow += 1
                OutlineBox(sheet, tagRow, catCol, tagRow + len(block.data[cat][tag]) - 1, catCol + 1)
                tagRow += len(block.data[cat][tag])
    # endregion

    # region write the review
        # write the review
        reviewRow = startRow
        for key in block.review:
            reviewCol = timeLine_colNum + 1
            for i in range(len(block.review[key])):
                sheet.cell(row=reviewRow, column=reviewCol + i, value=block.review[key][i])
                if weeklyInfoName[i] == "만족도":
                    conditionalFormatCell.append(chr(reviewCol + i - 1 + 65) + str(reviewRow))
            reviewRow += 1
    # endregion
        
    # region draw line for week separation
        # draw line on top of first week, while maintaining other part of border
        style = 'double'
        if block.date[-3] == "1":
            style = 'medium'
        for i in range(1, maxCol):
            brder = sheet.cell(row=startRow, column=i).border
            sheet.cell(row=startRow, column=i).border = oxl.styles.Border(top=oxl.styles.Side(style=style), bottom = brder.bottom, left = brder.left, right = brder.right)
        
        startRow += block.GetLength() # move to the next row for the next block
    # endregion

    # region Add month review        
        stackable = ['잘한 점', '개선점', '만족도']
        unStackable = ['목표', '계획', '성과', '이벤트']
        stack =[]
        singles = []
        # Add month review if it is the last block
        if (t == len(TimeBlocks.TimeBlocks) - 1) or (block.date[:8] != TimeBlocks.TimeBlocks[t+1].date[:8]) :
            for i in range(len(monthlyReview)):
                # find fitting month review
                if monthlyReview[i][0] == block.date[:9]:
                    MergeWriteCells(sheet, startRow, 1, startRow+1, 1, "월말 평가")
                    monthReviewCol = 2
                    for j in range(1,len(monthlyReview[i])):
                        colName = "목록"
                        if len(monthlyCategory) > j-1:
                            colName = monthlyCategory[j-1]
                        if colName in stackable:
                            stack =[]
                        else:
                            MergeWriteCells(sheet, startRow, monthReviewCol, startRow, monthReviewCol+1, colName, color = ColorPalette[j*3 + 5])
                            MergeWriteCells(sheet, startRow+1, monthReviewCol, startRow+1, monthReviewCol+1, monthlyReview[i][j])
                        monthReviewCol += 2
                    for j in range(1, maxCol):
                        sheet.cell(row=startRow, column=j).border = oxl.styles.Border(top=oxl.styles.Side(style='double'))
                    startRow += 2
    # endregion

    # region conditional format
    for cell in conditionalFormatCell:
        sheet.conditional_formatting.add(cell, DataBarRule(start_type="num", start_value=0, end_type="num", end_value=5, color=ColorPalette[-1], showValue="none"))
    
    # draw line for last row
    for i in range(1, maxCol):
        brder = sheet.cell(row=startRow-1, column=i).border
        sheet.cell(row=startRow-1, column=i).border = oxl.styles.Border(bottom=oxl.styles.Side(style='medium'), top = brder.top, left = brder.left, right = brder.right)

    WriteFirstRow(sheet, categoires, doIncludeWeeklyReview, weeklyInfoName)

    # width 32 16
    for i in range(timeLine_colNum):
        if i % 2 == 1:
            sheet.column_dimensions[chr(i+65)].width = 32
    sheet.column_dimensions[chr(timeLine_colNum - 1+65)].width = 16
    if doIncludeWeeklyReview:
        for i in range(len(weeklyInfoName)):
            sheet.column_dimensions[chr(timeLine_colNum + 1 + i + 65)].width = 24

    # set warp_text to true to all cells
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    lwb.remove(lwb["Sheet"]) # delete sheet name "Sheet"

    lwb.save(GetFileNames(person, dateRange, doIncludeMonthlyReview, doIncludeWeeklyReview, doFilterByDate))
    return

def WriteReviews(sheet, reviews, startRow, startCol,  reviewCategory = [], stackable = ['잘한 점', '개선점', '만족도'], unStackable = ['목표', '계획', '성과', '이벤트']):
    endRow = startRow
    endCol = startCol

    for i in range(len(reviews)):
        for j in range(len(reviews[i])):
            sheet.cell(row=startRow + i, column=startCol + j, value=reviews[i][j])
            if reviewCategory[j] == "만족도":
                conditionalFormatCell.append(chr(startCol + j - 1 + 65) + str(startRow + i))
    return endRow, endCol

def WriteFirstRow(sheet, categoires, weeklies = False, weeklyCategory = []):
    for i in range(len(categoires)):
        MergeWriteCells(sheet, 1, i*2 + 1, 1, i*2 + 2, categoires[i], color = ColorPalette[i*3 + 5])
    sheet.cell(row=1, column=len(categoires)*2 + 1, value="날짜")
    sheet.cell(row=1, column=len(categoires)*2 + 1).fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    if weeklies:
        for i in range(len(weeklyCategory)):
            sheet.cell(row=1, column=len(categoires)*2 + 2 + i, value=weeklyCategory[i])
            sheet.cell(row=1, column=len(categoires)*2 + 2 + i).fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    return

def GetFileNames(person = "김다은", dateRange = ["2024년 01월 1주차", "2024년 03월 1주차"], doIncludeMonthlyReview = True, doIncludeWeeklyReview = True, doFilterByDate = True):
    dateRangeStr = ""
    if doFilterByDate:
        dateRangeStr = "_("
        dateRangeStr += dateRange[0][:4] + dateRange[0][6:8] + dateRange[0][-3] + "~" + dateRange[1][:4] + dateRange[1][6:8] + dateRange[1][-3]
    if doIncludeMonthlyReview:
        if dateRangeStr == "":
            dateRangeStr += "_("
        dateRangeStr += "월말평가포함"
    if doIncludeWeeklyReview:
        if dateRangeStr == "":
            dateRangeStr += "_("
        dateRangeStr += "주간평가포함"
    if dateRangeStr != "":
        dateRangeStr += ")"
    file_name = "TimeLine_"+ person + dateRangeStr + ".xlsx"
    return file_name

def FilterDataByPerson(totalData, person):
    data = []

    personCol = 0
    for i in range(len(totalData[0])):
        if totalData[0][i] == "이름":
            personCol = i
            break
    
    data.append(totalData[0])
    for i in range(2, len(totalData)):
        if totalData[i][personCol] == person:
            data.append(totalData[i])
    return data

def FormatDate(data, date_colNum = 0, monthOnly = False):
    for i in range(1, len(data)):
        data[i][date_colNum] = data[i][date_colNum].strftime("%Y-%m-%d")
        if monthOnly:
            data[i][date_colNum] = data[i][date_colNum][:4] + "년 " + data[i][date_colNum][5:7] + "월"
        else:
            data[i][date_colNum] = data[i][date_colNum][:4] + "년 " + data[i][date_colNum][5:7] + "월 " + data[i][date_colNum][-1] + "주차"
    return data

def MergeWriteCells(sheet, start_row, start_col, end_row, end_col, value, color = None):
    sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    sheet.cell(row=start_row, column=start_col, value=value)
    sheet.cell(row=start_row, column=start_col).alignment = Alignment(horizontal="center", vertical="center")    # align the text to the center
    if color != None:
        sheet.cell(row=start_row, column=start_col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

def OutlineBox(sheet, start_row, start_col, end_row, end_col, style='thin'):
    if start_row == end_row and start_col == end_col:
        sheet.cell(row=start_row, column=start_col).border = oxl.styles.Border(top=oxl.styles.Side(style), bottom=oxl.styles.Side(style), left=oxl.styles.Side(style), right=oxl.styles.Side(style))
        return
    if start_row == end_row:
        for j in range(start_col, end_col+1):
            cell = sheet.cell(row=start_row, column=j)
            if j == start_col:
                cell.border = oxl.styles.Border(top=oxl.styles.Side(style), left=oxl.styles.Side(style), bottom=oxl.styles.Side(style))
            elif j == end_col:
                cell.border = oxl.styles.Border(top=oxl.styles.Side(style), right=oxl.styles.Side(style), bottom=oxl.styles.Side(style))
            else:
                cell.border = oxl.styles.Border(top=oxl.styles.Side(style), bottom=oxl.styles.Side(style))
        return
    if start_col == end_col:
        for i in range(start_row, end_row+1):
            cell = sheet.cell(row=i, column=start_col)
            if i == start_row:
                cell.border = oxl.styles.Border(top=oxl.styles.Side(style), left=oxl.styles.Side(style), right=oxl.styles.Side(style))
            elif i == end_row:
                cell.border = oxl.styles.Border(bottom=oxl.styles.Side(style), left=oxl.styles.Side(style), right=oxl.styles.Side(style))
            else:
                cell.border = oxl.styles.Border(left=oxl.styles.Side(style), right=oxl.styles.Side(style))
        return
    for i in range(start_row, end_row+1):
        for j in range(start_col, end_col+1):
            cell = sheet.cell(row=i, column=j)
            if i == start_row:
                if j == start_col:
                    cell.border = oxl.styles.Border(top=oxl.styles.Side(style), left=oxl.styles.Side(style))
                elif j == end_col:
                    cell.border = oxl.styles.Border(top=oxl.styles.Side(style), right=oxl.styles.Side(style))
                else:
                    cell.border = oxl.styles.Border(top=oxl.styles.Side(style))
            elif i == end_row:
                if j == start_col:
                    cell.border = oxl.styles.Border(bottom=oxl.styles.Side(style), left=oxl.styles.Side(style))
                elif j == end_col:
                    cell.border = oxl.styles.Border(bottom=oxl.styles.Side(style), right=oxl.styles.Side(style))
                else:
                    cell.border = oxl.styles.Border(bottom=oxl.styles.Side(style))
            else:
                if j == start_col:
                    cell.border = oxl.styles.Border(left=oxl.styles.Side(style))
                elif j == end_col:
                    cell.border = oxl.styles.Border(right=oxl.styles.Side(style))

def GetDateColor(date, dateColorPalette = DateColorPalette):
    month = int(date[6:8])
    weekNum = int(date[-3])
    return dateColorPalette[month-1][weekNum-1]

def GetTagColor(tag, colorPalette = ColorPalette, tagColorPalette = {}):
    if tag not in tagColorPalette:
        tagColorPalette[tag] = colorPalette[len(tagColorPalette)]
    return tagColorPalette[tag]

def RemoveNoneType(data):
    for i in range(len(data)):
        for j in range(len(data[i])):
            if data[i][j] == None:
                data[i][j] = "None"
    return data

def GetData(sheet, person, dateRange = None, excludingCols = ["이름", "etc"], ):
    # get the data
    total_data = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        total_data.append([cell.value for cell in row])

    data = FilterDataByPerson(total_data, person)
    data = FormatDate(data)
    data = RemoveNoneType(data)

    # exclude the columns that are not needed (DEFAULT: only left with date, category, tag, division, content)
    exColNums = []
    for i in range(len(excludingCols)):
        exCol = 0
        for j in range(len(data[0])):
            if data[0][j] == excludingCols[i]:
                exCol = j
                break
        exColNums.append(exCol)

    for i in range(len(excludingCols)):
        for j in range(len(data)):
            data[j].remove(data[j][exColNums[i]])
    
    # sort the data by date, category, tag
    data.sort(key=lambda x: (x[0], x[1], x[2]))

    # remove data that is not in the date range
    dateCol = 0
    for i in range(len(data[0])):
        if data[0][i] == "날짜":
            dateCol = i
            break

    if dateRange != None:
        data = [d for d in data if (d[dateCol] >= dateRange[0]) and (d[dateCol] <= dateRange[1])]
    
    return data

def RemoveUntouchedCategory(data, categories):
    removeCategory = []
    for i in range(len(categories)):
        found = False
        for j in range(len(data)):
            if data[j][1] == categories[i]:
                found = True
                break
        if not found:
            removeCategory.append(categories[i])

    for i in range(len(removeCategory)):
        categories.remove(removeCategory[i])
    return categories

def GetReview(sheet, person, removeCols = [1], monthOnly = False):
    total_review = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        total_review.append([cell.value for cell in row])

    review = FilterDataByPerson(total_review, person)
    review = FormatDate(review, monthOnly=monthOnly)
    for i in range(len(review)):
        for j in range(len(removeCols)):
            review[i].remove(review[i][removeCols[j]])
    return review

def GetMaxCol(TimeBlocks, categoires, review, monthlyReview):
    # check if there is any review included in the TimeBlocks
    noWeekReview = True
    for i in range(len(TimeBlocks.TimeBlocks)):
        if TimeBlocks.TimeBlocks[i].review != {}:
            noWeekReview = False
            break

    maxCol = 0
    timeLine_colNum = len(categoires) * 2 + 1
    if noWeekReview:
        maxCol = timeLine_colNum + 1
    else:
        maxCol = timeLine_colNum + len(review[0]) + 1

    # check if monthly review is longer than maxCol + check if there is any monthly review included in dates
    if (monthlyReview != [])and (len(monthlyReview[0]) > maxCol):
        for i in range(len(monthlyReview)):
            if monthlyReview[i][0] in [block.date[:9] for block in TimeBlocks.TimeBlocks]:
                maxCol = len(monthlyReview[i])*2
    return maxCol

def GetCategories(sheet, name = "Category"):
    # find 'name' coloumn in first row
    colNum = 0
    for cell in sheet[1]:
        if cell.value == name:
            colNum = cell.col_idx
            break
    
    # get the categories
    categories = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=colNum, max_col=colNum):
        categories.append(row[0].value)
    return categories

def GetProcessedData(sheet, categories, person, dateRange = None, excludingCols = ["이름", "etc"], categories = None):
    data = GetData(sheet, person, dateRange, excludingCols)
    if categories == None:
        categories = GetCategories(sheet)
    RemoveUntouchedCategory(data, categories)
    return data