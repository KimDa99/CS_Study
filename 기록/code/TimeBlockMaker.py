import openpyxl as oxl
from openpyxl import load_workbook
import numpy as np
from datetime import datetime
import TimeBlock

class TimeBlockMaker:
    person = ""
    doFilterByDate = False
    doIncludeWeeklyReview = True
    doIncludeMonthlyReview = True
    dateRange = []

    total_sheetName = "전체 한 일"
    weeklyCheck_sheetName = "주간 점검"
    monthlyCheck_sheetName = "월 점검"
    categories_sheetName = "categories"

    TimeBlocks = None
    MonthBlocks = None

    def __init__(self, person, doFilterByDate = False, doIncludeMonthlyReview=True, doIncludeWeeklyReview=True, dateRange = []):
        self.person = person
        self.doFilterByDate = doFilterByDate
        self.doIncludeMonthlyReview = doIncludeMonthlyReview
        self.doIncludeWeeklyReview = doIncludeWeeklyReview
        self.dateRange = dateRange
    
    def LoadTimeBlocks(self, fileName = "Data.xlsx", totalSheetName="전체 한 일", weeklyCheckSheetName="주간 점검", monthlyCheckSheetName="월 점검", categoriesSheetName="categories"):
        # create a workbook object
        wb = oxl.Workbook()
        lwb = oxl.load_workbook(fileName)

        # get the datas
        data = self.GetData(lwb)
        weeklyReview = self.GetWeeklyReview(lwb)
        monthlyReview = self.GetMonthlyReview(lwb)

        data = self.FormatDate(data)
        weeklyReview = self.FormatDate(weeklyReview)
        monthlyReview = self.FormatDate(monthlyReview, monthOnly=True)

        # create a time block list
        self.TimeBlocks = TimeBlock.TimeBlockList(data, reviews=weeklyReview)
        self.MonthBlocks = TimeBlock.MonthBlockList(data, self.TimeBlocks.TimeBlocks, monthlyReview)

        wb.close()
    
    def GetData(self, lwb):
        total_data = []
        
        sheet = lwb[self.total_sheetName]
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            total_data.append(row)
        
        data = self.FilterByName(total_data)

        return data
    
    def GetWeeklyReview(self, lwb):
        weekly_data = []
        
        sheet = lwb[self.weeklyCheck_sheetName]
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            weekly_data.append(row)
        
        data = self.FilterByName(weekly_data)

        return data
    
    def GetMonthlyReview(self, lwb):
        monthly_data = []
        
        sheet = lwb[self.monthlyCheck_sheetName]
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            monthly_data.append(row)
        
        data = self.FilterByName(monthly_data)

        return data

    def FilterByName(self, data, colNameStirng = "이름"):
        filtered = []

        nameIndex = -1
        for i in range(len(data[0])):
            if data[0][i] == colNameStirng:
                nameIndex = i
                break
        
        if nameIndex == -1:
            print("No name column found with the given data")
            return data
        
        for row in data:
            if row[nameIndex] == self.person:
                filtered.append(row)

        return filtered

    def FormatDate(self, data, monthOnly = False):
        date_colNum = -1

        # get from the first row
        for i in range (len(data[0])):
            if data[0][i] == "날짜":
                date_colNum = i
                break

        for i in range(1, len(data)):
            data[i][date_colNum] = data[i][date_colNum].strftime("%Y-%m-%d")
            if monthOnly:
                data[i][date_colNum] = data[i][date_colNum][:4] + "년 " + data[i][date_colNum][5:7] + "월"
            else:
                data[i][date_colNum] = data[i][date_colNum][:4] + "년 " + data[i][date_colNum][5:7] + "월 " + data[i][date_colNum][-1] + "주차"
        return data